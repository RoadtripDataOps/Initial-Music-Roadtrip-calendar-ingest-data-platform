import csv
import json
from collections.abc import Iterable, Mapping
from datetime import datetime, time
from io import BytesIO, StringIO

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.db.models import (
    ImportBatch,
    MasterCalendarSource,
    StagedCalendarSource,
    StagedEvent,
)
from app.services.event_dedupe_service import (
    NormalizedEventCandidate,
    SourceClaimInput,
    upsert_event_from_candidate,
)
from app.services.file_risk_service import (
    is_direct_image_url,
    is_full_url,
    is_social_url,
    score_calendar_source_rows,
    score_concert_event_rows,
)
from app.services.image_qa_service import (
    ImageCandidateInput,
    create_image_candidate,
    select_best_event_image,
)
from app.services.master_calendar_service import (
    CalendarSourcePayload,
    canonicalize_calendar_url,
    create_or_attach_master_calendar_source,
    get_master_by_hash,
    normalize_expected_category,
)
from app.services.risk_service import RiskAssessment, build_assessment
from app.services.security_service import neutralize_csv_formula
from app.services.venue_service import (
    VenueInput,
    ensure_event_venue,
    parse_optional_float,
)

CALENDAR_SOURCE_HEADERS = [
    "Organization Name",
    "Calendar Name",
    "Calendar URL",
    "Source Type",
    "Expected Category",
    "Venue Name",
    "City",
    "State",
    "Country",
    "Region / Market",
    "Contact Name",
    "Contact Email",
    "Crawl Frequency",
    "Authorization Confirmed",
    "Notes",
]

CONCERT_EVENT_HEADERS = [
    "Category",
    "Event Name",
    "Headliner",
    "Supporting Artists",
    "Start Date",
    "Start Time",
    "Timezone",
    "End Date",
    "End Time",
    "Doors Time",
    "Venue Name",
    "Venue Address",
    "City",
    "State",
    "Zip Code",
    "Country",
    "Latitude",
    "Longitude",
    "Event URL",
    "Tickets Link",
    "Description",
    "Price",
    "Age Restriction",
    "Phone",
    "Email",
    "Website",
    "Spotify URL",
    "YouTube URL",
    "Instagram",
    "Facebook",
    "X",
    "TikTok",
    "Main Image URL",
    "Additional Image URL(s)",
    "Source Event ID",
    "Notes",
]

URL_FIELDS = [
    "Event URL",
    "Tickets Link",
    "Website",
    "Spotify URL",
    "YouTube URL",
    "Instagram",
    "Facebook",
    "X",
    "TikTok",
]


class ImportValidationError(Exception):
    """Raised when an uploaded import file cannot be staged."""


def clean_cell(value: object | None) -> str:
    if value is None:
        return ""
    return str(value).strip()


def truthy(value: str) -> bool:
    return value.strip().lower() in {"true", "yes", "1", "y"}


def rows_from_csv(content: bytes) -> list[dict[str, str]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(text))
    if reader.fieldnames is None:
        raise ImportValidationError("Uploaded CSV is empty or missing headers.")
    return [{key: clean_cell(value) for key, value in row.items()} for row in reader]


def rows_from_xlsx(content: bytes) -> list[dict[str, str]]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ImportValidationError(
            "Uploaded XLSX could not be read; password-protected or invalid "
            "workbooks are not accepted.",
        ) from exc
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ImportValidationError("Uploaded XLSX is empty or missing headers.")
    headers = [clean_cell(value) for value in rows[0]]
    if not any(headers):
        raise ImportValidationError("Uploaded XLSX is empty or missing headers.")
    parsed_rows: list[dict[str, str]] = []
    for row in rows[1:]:
        parsed_rows.append(
            {
                headers[index]: clean_cell(value)
                for index, value in enumerate(row)
                if index < len(headers)
            }
        )
    return parsed_rows


def parse_upload_rows(
    filename: str,
    content: bytes,
    max_rows: int | None = None,
) -> tuple[str, list[dict[str, str]]]:
    if not content:
        raise ImportValidationError("Uploaded file is empty.")
    lowered = filename.lower()
    if lowered.endswith(".csv"):
        rows = rows_from_csv(content)
        file_type = "csv"
    elif lowered.endswith(".xlsx"):
        rows = rows_from_xlsx(content)
        file_type = "xlsx"
    else:
        raise ImportValidationError("Unsupported file type. Upload CSV or XLSX.")
    if max_rows is not None and len(rows) > max_rows:
        raise ImportValidationError("Uploaded file has too many rows.")
    return file_type, rows


def require_headers(rows: list[dict[str, str]], required_headers: list[str]) -> None:
    if not rows:
        raise ImportValidationError("Uploaded file contains no data rows.")
    headers = set(rows[0].keys())
    missing = [header for header in required_headers if header not in headers]
    if missing:
        raise ImportValidationError(f"Missing required headers: {', '.join(missing)}")


def csv_template(headers: list[str]) -> str:
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    return output.getvalue()


def xlsx_template(headers: list[str]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def json_list(values: Iterable[str]) -> str:
    return json.dumps(sorted(set(values)), ensure_ascii=True)


def row_json(row: Mapping[str, str]) -> str:
    preview_safe = {
        key: neutralize_csv_formula(value) for key, value in dict(row).items()
    }
    return json.dumps(preview_safe, ensure_ascii=True, sort_keys=True)


def validation_status(errors: list[str]) -> str:
    return "invalid" if errors else "valid"


def row_assessment(flags: list[str]) -> RiskAssessment:
    return build_assessment(len(set(flags)) * 10, flags)


def combine_assessments(*assessments: RiskAssessment | None) -> RiskAssessment:
    score = 0
    flags: list[str] = []
    for assessment in assessments:
        if assessment is None:
            continue
        score += assessment.risk_score
        flags.extend(assessment.risk_flags)
    return build_assessment(score, flags)


def validate_concert_row(row: Mapping[str, str]) -> tuple[list[str], RiskAssessment]:
    errors: list[str] = []
    category = clean_cell(row.get("Category")) or "Concert"
    if category != "Concert":
        errors.append("non_concert_category")

    required = {
        "Event Name": "event_name_missing",
        "Headliner": "headliner_missing",
        "Start Date": "start_date_missing",
        "Timezone": "timezone_missing",
        "Venue Name": "venue_name_missing",
        "City": "city_missing",
        "State": "state_missing",
    }
    for field_name, flag in required.items():
        if not clean_cell(row.get(field_name)):
            errors.append(flag)

    if not clean_cell(row.get("Event URL")) and not clean_cell(row.get("Tickets Link")):
        errors.append("event_or_ticket_url_missing")
    if not clean_cell(row.get("Venue Address")) and not (
        clean_cell(row.get("Latitude")) and clean_cell(row.get("Longitude"))
    ):
        errors.append("venue_address_or_coordinates_missing")

    for field_name in URL_FIELDS:
        value = clean_cell(row.get(field_name))
        if value and not is_full_url(value):
            errors.append(f"{field_name.lower().replace(' ', '_')}_not_full_url")

    main_image = clean_cell(row.get("Main Image URL"))
    if main_image:
        if is_social_url(main_image):
            errors.append("main_image_social_media_url")
        elif not is_direct_image_url(main_image):
            errors.append("main_image_not_direct_public_image")

    additional_images = clean_cell(row.get("Additional Image URL(s)"))
    image_urls = [
        part.strip() for part in additional_images.split("$") if part.strip()
    ]
    for image_url in image_urls:
        if is_social_url(image_url):
            errors.append("additional_image_social_media_url")
        elif not is_direct_image_url(image_url):
            errors.append("additional_image_not_direct_public_image")

    risk = score_concert_event_rows([row])
    return sorted(set(errors + risk.risk_flags)), risk


def validate_calendar_source_row(
    row: Mapping[str, str],
    seen_hashes: set[str],
    session: Session,
) -> tuple[list[str], RiskAssessment, str | None, str | None, str, int | None]:
    errors: list[str] = []
    organization = clean_cell(row.get("Organization Name"))
    calendar_url = clean_cell(row.get("Calendar URL"))
    contact_email = clean_cell(row.get("Contact Email"))
    expected_category = normalize_expected_category(
        clean_cell(row.get("Expected Category"))
    )

    if not organization:
        errors.append("organization_name_missing")
    if not calendar_url:
        errors.append("calendar_url_missing")
    if not contact_email:
        errors.append("contact_email_missing")
    if not truthy(clean_cell(row.get("Authorization Confirmed"))):
        errors.append("authorization_missing")
    if expected_category != "Concert":
        errors.append("expected_category_not_concert")

    canonical_url: str | None = None
    url_hash: str | None = None
    dedupe_status = "new"
    existing_id: int | None = None
    if calendar_url:
        canonical_url, url_hash = canonicalize_calendar_url(calendar_url)
        existing = get_master_by_hash(session, url_hash)
        if url_hash in seen_hashes:
            dedupe_status = "duplicate_within_file"
        elif existing is not None:
            dedupe_status = "duplicate_existing"
            existing_id = existing.id
        seen_hashes.add(url_hash)

    risk = score_calendar_source_rows([row])
    return (
        sorted(set(errors + risk.risk_flags)),
        risk,
        canonical_url,
        url_hash,
        dedupe_status,
        existing_id,
    )


def create_import_batch(
    session: Session,
    submission_type: str,
    organization_name: str,
    contact_name: str,
    contact_email: str,
    filename: str,
    file_type: str,
    rows: list[dict[str, str]],
    assessment: RiskAssessment,
    notes: str | None,
) -> ImportBatch:
    batch = ImportBatch(
        submission_type=submission_type,
        organization_name=organization_name,
        contact_name=contact_name,
        contact_email=contact_email,
        original_filename=filename,
        file_type=file_type,
        status="staged",
        review_status="pending_review"
        if assessment.risk_level not in {"high", "blocked"}
        else "quarantined",
        risk_score=assessment.risk_score,
        risk_level=assessment.risk_level,
        risk_flags_json=assessment.risk_flags_json,
        notes=notes,
        valid_row_count=0,
        invalid_row_count=0,
        duplicate_row_count=0,
    )
    session.add(batch)
    session.flush()
    return batch


def stage_concert_events_upload(
    session: Session,
    organization_name: str,
    contact_name: str,
    contact_email: str,
    filename: str,
    content: bytes,
    notes: str | None = None,
    gate_assessment: RiskAssessment | None = None,
    max_rows: int | None = None,
) -> ImportBatch:
    file_type, rows = parse_upload_rows(filename, content, max_rows=max_rows)
    require_headers(rows, CONCERT_EVENT_HEADERS)
    batch_assessment = combine_assessments(
        score_concert_event_rows(rows),
        gate_assessment,
    )
    batch = create_import_batch(
        session,
        "concert_events_file",
        organization_name,
        contact_name,
        contact_email,
        filename,
        file_type,
        rows,
        batch_assessment,
        notes,
    )

    valid_count = 0
    invalid_count = 0
    for index, row in enumerate(rows, start=2):
        errors, risk = validate_concert_row(row)
        category = clean_cell(row.get("Category")) or "Concert"
        status = validation_status(errors)
        if status == "valid":
            valid_count += 1
        else:
            invalid_count += 1
        session.add(
            StagedEvent(
                import_batch_id=batch.id,
                row_number=index,
                validation_status=status,
                validation_errors=json_list(errors),
                risk_score=risk.risk_score,
                risk_level=risk.risk_level,
                risk_flags_json=risk.risk_flags_json,
                category=category,
                event_name=clean_cell(row.get("Event Name")) or None,
                headliner=clean_cell(row.get("Headliner")) or None,
                supporting_artists=clean_cell(row.get("Supporting Artists")) or None,
                start_date=clean_cell(row.get("Start Date")) or None,
                start_time=clean_cell(row.get("Start Time")) or None,
                timezone=clean_cell(row.get("Timezone")) or None,
                end_date=clean_cell(row.get("End Date")) or None,
                end_time=clean_cell(row.get("End Time")) or None,
                doors_time=clean_cell(row.get("Doors Time")) or None,
                venue_name=clean_cell(row.get("Venue Name")) or None,
                venue_address=clean_cell(row.get("Venue Address")) or None,
                city=clean_cell(row.get("City")) or None,
                state=clean_cell(row.get("State")) or None,
                zip_code=clean_cell(row.get("Zip Code")) or None,
                country=clean_cell(row.get("Country")) or None,
                latitude=clean_cell(row.get("Latitude")) or None,
                longitude=clean_cell(row.get("Longitude")) or None,
                event_url=clean_cell(row.get("Event URL")) or None,
                tickets_link=clean_cell(row.get("Tickets Link")) or None,
                description=clean_cell(row.get("Description")) or None,
                price=clean_cell(row.get("Price")) or None,
                age_restriction=clean_cell(row.get("Age Restriction")) or None,
                phone=clean_cell(row.get("Phone")) or None,
                email=clean_cell(row.get("Email")) or None,
                website=clean_cell(row.get("Website")) or None,
                spotify_url=clean_cell(row.get("Spotify URL")) or None,
                youtube_url=clean_cell(row.get("YouTube URL")) or None,
                instagram=clean_cell(row.get("Instagram")) or None,
                facebook=clean_cell(row.get("Facebook")) or None,
                x_url=clean_cell(row.get("X")) or None,
                tiktok=clean_cell(row.get("TikTok")) or None,
                main_image_url=clean_cell(row.get("Main Image URL")) or None,
                additional_image_urls=(
                    clean_cell(row.get("Additional Image URL(s)")) or None
                ),
                source_event_id=clean_cell(row.get("Source Event ID")) or None,
                notes=clean_cell(row.get("Notes")) or None,
                raw_row_json=row_json(row),
            )
        )

    batch.valid_row_count = valid_count
    batch.invalid_row_count = invalid_count
    session.add(batch)
    session.commit()
    session.refresh(batch)
    return batch


def stage_calendar_sources_upload(
    session: Session,
    organization_name: str,
    contact_name: str,
    contact_email: str,
    filename: str,
    content: bytes,
    notes: str | None = None,
    gate_assessment: RiskAssessment | None = None,
    max_rows: int | None = None,
) -> ImportBatch:
    file_type, rows = parse_upload_rows(filename, content, max_rows=max_rows)
    require_headers(rows, CALENDAR_SOURCE_HEADERS)
    batch_assessment = combine_assessments(
        score_calendar_source_rows(rows),
        gate_assessment,
    )
    batch = create_import_batch(
        session,
        "calendar_sources_file",
        organization_name,
        contact_name,
        contact_email,
        filename,
        file_type,
        rows,
        batch_assessment,
        notes,
    )

    seen_hashes: set[str] = set()
    valid_count = 0
    invalid_count = 0
    duplicate_count = 0
    for index, row in enumerate(rows, start=2):
        errors, risk, canonical, url_hash, dedupe_status, existing_id = (
            validate_calendar_source_row(row, seen_hashes, session)
        )
        if dedupe_status != "new":
            duplicate_count += 1
        status = validation_status(errors)
        if status == "valid":
            valid_count += 1
        else:
            invalid_count += 1
        expected_category = normalize_expected_category(
            clean_cell(row.get("Expected Category"))
        )
        source_type = clean_cell(row.get("Source Type")) or "unknown"
        session.add(
            StagedCalendarSource(
                import_batch_id=batch.id,
                row_number=index,
                validation_status=status,
                validation_errors=json_list(errors),
                risk_score=risk.risk_score,
                risk_level=risk.risk_level,
                risk_flags_json=risk.risk_flags_json,
                organization_name=clean_cell(row.get("Organization Name")) or None,
                calendar_name=clean_cell(row.get("Calendar Name")) or None,
                calendar_url=clean_cell(row.get("Calendar URL")) or None,
                canonical_url=canonical,
                canonical_url_hash=url_hash,
                source_type=source_type,
                expected_category=expected_category,
                venue_name=clean_cell(row.get("Venue Name")) or None,
                city=clean_cell(row.get("City")) or None,
                state=clean_cell(row.get("State")) or None,
                country=clean_cell(row.get("Country")) or None,
                region_or_market=clean_cell(row.get("Region / Market")) or None,
                contact_name=clean_cell(row.get("Contact Name")) or None,
                contact_email=clean_cell(row.get("Contact Email")) or None,
                crawl_frequency=clean_cell(row.get("Crawl Frequency")) or None,
                authorization_confirmed=truthy(
                    clean_cell(row.get("Authorization Confirmed"))
                ),
                dedupe_status=dedupe_status,
                existing_master_calendar_source_id=existing_id,
                notes=clean_cell(row.get("Notes")) or None,
                raw_row_json=row_json(row),
            )
        )

    batch.valid_row_count = valid_count
    batch.invalid_row_count = invalid_count
    batch.duplicate_row_count = duplicate_count
    session.add(batch)
    session.commit()
    session.refresh(batch)
    return batch


def list_import_batches(session: Session) -> list[ImportBatch]:
    statement = select(ImportBatch).order_by(ImportBatch.created_at.desc())
    return list(session.scalars(statement).all())


def get_import_batch(session: Session, batch_id: int) -> ImportBatch | None:
    return session.get(ImportBatch, batch_id)


def staged_events_for_batch(session: Session, batch_id: int) -> list[StagedEvent]:
    statement = (
        select(StagedEvent)
        .where(StagedEvent.import_batch_id == batch_id)
        .order_by(StagedEvent.row_number.asc())
    )
    return list(session.scalars(statement).all())


def staged_sources_for_batch(
    session: Session,
    batch_id: int,
) -> list[StagedCalendarSource]:
    statement = (
        select(StagedCalendarSource)
        .where(StagedCalendarSource.import_batch_id == batch_id)
        .order_by(StagedCalendarSource.row_number.asc())
    )
    return list(session.scalars(statement).all())


def parse_event_datetime(row: StagedEvent) -> datetime:
    start_date = row.start_date or "1970-01-01"
    start_time = row.start_time or "00:00"
    return datetime.combine(
        datetime.fromisoformat(start_date).date(),
        time.fromisoformat(start_time),
    )


def parse_optional_event_datetime(
    date_value: str | None,
    time_value: str | None,
) -> datetime | None:
    if not date_value:
        return None
    try:
        return datetime.combine(
            datetime.fromisoformat(date_value).date(),
            time.fromisoformat(time_value or "00:00"),
        )
    except ValueError:
        return None


def approve_valid_staged_events(session: Session, batch_id: int) -> int:
    batch = session.get(ImportBatch, batch_id)
    if batch is None or batch.review_status in {"quarantined", "blocked", "rejected"}:
        return 0

    created = 0
    updated = 0
    duplicates = 0
    claims = 0
    rejected = 0
    quarantined = 0
    for row in staged_events_for_batch(session, batch_id):
        if row.validation_status != "valid":
            rejected += 1
            continue
        if row.risk_level in {"high", "blocked"}:
            quarantined += 1
            continue
        venue = ensure_event_venue(
            session,
            VenueInput(
                display_name=row.venue_name or "Unknown Venue",
                address=row.venue_address,
                city=row.city,
                state=row.state,
                zip_code=row.zip_code,
                country=row.country,
                latitude=parse_optional_float(row.latitude),
                longitude=parse_optional_float(row.longitude),
                website=row.website,
                phone=row.phone,
                description=None,
                main_image_url=row.main_image_url,
                additional_image_urls=row.additional_image_urls,
            ),
        )
        raw_json = row.raw_row_json or "{}"
        normalized = NormalizedEventCandidate(
            event_venue_id=venue.id,
            import_batch_id=batch_id,
            category="Concert",
            record_type="event",
            source_type="file_upload",
            ingestion_provider="file_upload",
            title=row.event_name or "Untitled Concert",
            headliner=row.headliner,
            supporting_artists=row.supporting_artists,
            genre=None,
            description=row.description,
            start_datetime=parse_event_datetime(row),
            end_datetime=parse_optional_event_datetime(row.end_date, row.end_time),
            timezone=row.timezone,
            location_text=row.venue_name or row.venue_address,
            source_url=row.event_url,
            tickets_link=row.tickets_link,
            price=row.price,
            age_restriction=row.age_restriction,
            doors_time=row.doors_time,
            main_image_url=row.main_image_url,
            additional_image_urls=row.additional_image_urls,
            spotify_url=row.spotify_url,
            source_event_id=row.source_event_id,
            raw_event_json=raw_json,
            source_chain_json=json.dumps(
                [
                    {
                        "role": "file_upload",
                        "source": "concert_events_file",
                        "identifier": str(row.id),
                    }
                ],
                ensure_ascii=True,
                sort_keys=True,
            ),
        )
        result = upsert_event_from_candidate(
            session,
            normalized,
            SourceClaimInput(
                source_type="file_upload",
                ingestion_provider="file_upload",
                provider_record_id=str(row.id),
                source_record_id=(
                    row.source_event_id or f"batch-{batch_id}-row-{row.row_number}"
                ),
                source_url=row.event_url or row.tickets_link,
                source_name=batch.organization_name,
                import_batch_id=batch_id,
                raw_payload_json=raw_json,
                normalized_payload_json=raw_json,
                field_values={
                    "event_name": row.event_name,
                    "headliner": row.headliner,
                    "start_datetime": parse_event_datetime(row),
                    "venue_name": row.venue_name,
                },
                source_chain_json=normalized.source_chain_json or "[]",
            ),
        )
        claims += 1
        if result.action == "created":
            created += 1
        elif result.action == "updated":
            updated += 1
        elif result.action == "duplicate_candidate":
            duplicates += 1
        if row.main_image_url:
            create_image_candidate(
                session,
                ImageCandidateInput(
                    event_id=result.event.id,
                    source_type="upload",
                    source_provider="concert_events_file",
                    source_url=row.event_url or row.tickets_link,
                    image_url=row.main_image_url,
                    image_role="event_provider",
                    clearance_status="needs_approval",
                ),
                commit=False,
            )
            select_best_event_image(session, result.event.id, commit=False)
    batch.review_status = "approved"
    batch.status = "approved"
    batch.events_created_count = created
    batch.events_updated_count = updated
    batch.duplicate_candidate_count = duplicates
    batch.source_claims_created_count = claims
    batch.rows_rejected_count = rejected
    batch.rows_quarantined_count = quarantined
    session.add(batch)
    session.commit()
    return created


def approve_valid_staged_calendar_sources(session: Session, batch_id: int) -> int:
    batch = session.get(ImportBatch, batch_id)
    if batch is None or batch.review_status in {"quarantined", "blocked", "rejected"}:
        return 0

    created = 0
    hash_to_master: dict[str, MasterCalendarSource] = {}
    for row in staged_sources_for_batch(session, batch_id):
        if row.validation_status != "valid" or row.risk_level in {"high", "blocked"}:
            continue
        if not row.calendar_url or not row.canonical_url_hash:
            continue
        existing = get_master_by_hash(session, row.canonical_url_hash)
        if existing is None and row.canonical_url_hash in hash_to_master:
            existing = hash_to_master[row.canonical_url_hash]

        assessment = RiskAssessment(
            risk_score=row.risk_score,
            risk_level=row.risk_level,
            risk_flags=row.risk_flags,
        )
        payload = CalendarSourcePayload(
            organization_name=row.organization_name or batch.organization_name or "",
            contact_name=row.contact_name,
            contact_email=row.contact_email or batch.contact_email or "",
            calendar_name=row.calendar_name,
            calendar_url=row.calendar_url,
            source_type=row.source_type,
            expected_category=row.expected_category,
            venue_name=row.venue_name,
            city=row.city,
            state=row.state,
            country=row.country,
            region_or_market=row.region_or_market,
            crawl_frequency=row.crawl_frequency,
            authorization_confirmed=row.authorization_confirmed,
            notes=row.notes,
            import_batch_id=batch_id,
            raw_row_json=row.raw_row_json,
        )
        master, _submission, was_created = create_or_attach_master_calendar_source(
            session,
            payload,
            assessment,
            "pending_review",
        )
        hash_to_master[row.canonical_url_hash] = master
        if was_created and existing is None:
            created += 1
    batch.review_status = "approved"
    batch.status = "approved"
    session.add(batch)
    session.commit()
    return created


def reject_or_quarantine_batch(
    session: Session,
    batch_id: int,
    action: str,
) -> ImportBatch | None:
    batch = session.get(ImportBatch, batch_id)
    if batch is None:
        return None
    if action == "reject":
        batch.review_status = "rejected"
        batch.status = "rejected"
    elif action == "quarantine":
        batch.review_status = "quarantined"
        batch.status = "quarantined"
    else:
        return None
    session.add(batch)
    session.commit()
    session.refresh(batch)
    return batch
